import argparse
import openpyxl
import re
import sys
import openpyxl.styles.builtins

from copy import copy
from itertools import groupby
from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import DEFAULT_FONT, Alignment, Border, Font, NamedStyle, Side


YEAR = '2025-2026'
artists = {
        'front':        ['Cadence MacDougall', '5th grade'],
        'front inside': ['Patrick Riehman', '3rd grade'],
        'back inside':  ['Olivia Stelle', '4th grade'],
        'back':         ['Mabel McCahill', '4th grade'],
        }

openpyxl.styles.builtins.hyperlink

def fix_name(s):
    # Mcdonald -> McDonald
    name = re.sub(r"Mc([a-z])", lambda m: "Mc" + m.group(1).upper(), s)

    # (Rj) -> (RJ).  Hopefully they do not put nicknames that aren't capitalized in parens
    name = re.sub(r"(\(.*?\))", lambda m: m.group(1).upper(), name)

    return name

class Student:
    def __init__(self, name, grade, teacher, guardians=None):
        # Strip any suffixes like I, II, III, etc since that was included in the class list but not the parent file
        self.name = re.sub(r", I+$", "", name.strip())
        self.title = fix_name(re.sub(r"(.+),\s+(.+)", r"\2 \1", self.name).title())
        self.index_name = fix_name(self.name.title())
        self.grade = grade
        self.teacher = teacher
        self.guardians = guardians

    def __repr__(self):
        return str(self)

    def __hash__(self):
        return hash((self.grade, self.name))

    def __eq__(self, other):
        a = (self.grade, self.name)
        b = (other.grade, other.name)
        return a == b

    def __lt__(self, other):
        return (self.grade, self.name) < (other.grade, other.name)

    def __str__(self):
        return f"{self.name} {self.title} - Grade {self.grade} - Teacher {self.teacher.title} - Guardians {self.guardians}"

    def address(self):
        if self.guardians:
            addresses = [g.address for g in self.guardians if g.address]
            if addresses:
                return next((a for a in addresses if a is not None), None)
            else:
                return None

        else:
            return None

    @staticmethod
    def parse_from_parent_file(row, has_phone, has_address, guardian_2_index):
        grade = Grade(row[1].value)

        guardians = [Guardian(
                name=row[4].value,
                email=row[5].value,
                phone=row[6].value if has_phone else None,
                address=row[7].value if has_address else None)]

        if len(row) > guardian_2_index:
            if row[guardian_2_index].value:
                name2 = row[guardian_2_index].value if row[guardian_2_index] else None
                email2 = row[guardian_2_index + 1].value if row[guardian_2_index + 1] else None
                phone2 = row[guardian_2_index + 2].value if row[guardian_2_index + 2] else None
                guardians.append(Guardian(
                    name=name2,
                    email=email2,
                    phone=phone2,
                    ))

        teacher = Teacher(name=row[2].value, grade=grade)

        return Student(
                name=row[0].value,
                grade=grade,
                teacher=teacher,
                guardians=guardians)

class Grade:
    def __init__(self, value):
        stripped = re.sub(r"(ST|ND|RD|TH|DG)", "", str(value)).strip()
        if stripped == 'K':
            self.grade = 'K'
            self.order = 0
        else:
            self.grade = int(float(stripped))
            self.order = self.grade

    def pretty(self):
        match self.order:
            case 0: return "Kindergarten"
            case 1: return "1st Grade"
            case 2: return "2nd Grade"
            case 3: return "3rd Grade"
            case 4: return "4th Grade"
            case 5: return "5th Grade"
            case _: raise ValueError("Invalid grade")

    def __lt__(self, other):
        return self.order < other.order

    def __hash__(self):
        return hash(self.order)

    def __repr__(self):
        return str(self)

    def __str__(self):
        return str(self.grade)

    def __eq__(self, other):
        return self.order == other.order

class Teacher:
    def __init__(self, name, grade):
        self.name = name
        if self.name == 'MORGAN EVANCIC':
            self.name = 'MORGAN BAETZ'
        self.grade = grade
        self.title = self.name.title()
        self.class_list_lookup = self.name.split(" ")[-1]
        self.students = []

    def __lt__(self, other):
        return (self.grade, self.name) < (other.grade, other.name)

    def __repr__(self):
        return str(self)

    def __str__(self):
        return f"{self.grade} - {self.title}"

    def __eq__(self, other):
        a = (self.grade, self.class_list_lookup)
        b = (other.grade, other.class_list_lookup)
        return a == b

    def add_students(self, students):
        self.students.extend(students)

class Class:
    def __init__(self, room, teacher, grade, students):
        self.room = re.sub(r"# ", "", room.title())
        self.teacher = teacher
        self.grade = grade
        self.students = students

    def title(self):
        return f"{self.grade.grade} {self.teacher.class_list_lookup}"

    def __repr__(self):
        return str(self)

    def __str__(self):
        return f"{self.grade} - {self.teacher} - {self.students}"

class Guardian:
    def __init__(self, name, email, phone=None, address=None):
        self.name = name
        self.email = email.lower()
        self.phone = phone
        self.address = address.title() if address else None
        if self.address:
            self.address = re.sub(r"Lombard, IL 60148", "", self.address, flags=re.IGNORECASE)

    def title(self):
        return fix_name(self.name.title())

    def phone_link(self):
        return f"https://call.ctrlq.org/1{self.phone.replace('-', '')}" if self.phone else ''

    def email_link(self):
        return f"mailto:{self.email}" if self.email else ''

    def __repr__(self):
        return str(self)

    def __str__(self):
        return f"{self.title()} {self.email} {self.phone} {self.address}"

class ParentParser:
    @staticmethod
    def parse_parent_students(parent_files):
        return [s for f in parent_files for s in ParentParser.__parse_parent_file(f)]

    @staticmethod
    def __parse_parent_file(f):
        wb = openpyxl.load_workbook(f)
        sheet = wb.active
        has_phone = sheet['G1'].value == 'Phone'
        has_address = sheet['H1'] and sheet['H1'].value == 'Address'
        if has_address:
            guardian_2_index = 10
        elif has_phone:
            guardian_2_index = 7
        else:
            guardian_2_index = 6

        return [Student.parse_from_parent_file(row, has_phone, has_address, guardian_2_index) for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)]

class ClassListParser:
    @staticmethod
    def parse_class(sheet):
        teacher_name = re.sub(r" \(.*\)$", "", sheet.cell(1, 1).value.upper().replace('TEACHER: ', '').strip())
        # Transform Kdg, 1st, 2nd, 3rd, 4th, 5th => K, 1, 2, 3, 4, 5
        grade = Grade(sheet.cell(1, 2).value)
        room = sheet.cell(1, 3).value

        if teacher_name != sheet.title:
            raise Exception(f"Expected teacher name {teacher_name} to match sheet title {sheet.title}")

        teacher = Teacher(teacher_name, grade)
        students = ClassListParser.parse_students(teacher, sheet)

        return Class(room, teacher, grade, students)

    @staticmethod
    def parse_students(teacher, sheet):
        students = []
        for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if row[0] is None or row[0].value is None or "total" in row[0].value.lower():
                break

            s = Student(name=row[0].value, grade=teacher.grade, teacher=teacher)
            students.append(s)
        return students

    @staticmethod
    def parse_lists(class_list):
        wb = openpyxl.load_workbook(class_list)

        teachers = []
        for sheet in wb.worksheets:
            if sheet.title.startswith("Sheet"):
                continue

            teacher = ClassListParser.parse_class(sheet)
            teachers.append(teacher)

        return teachers


class TextOutput:

    blank_guardian = Guardian(name='', email='', phone='', address='')

    def print_class(self, cls):
        print(cls.teacher.title)
        print(f"{cls.grade.pretty()} - {cls.room}")

        for s in cls.students:
            address = s.address() if s.address() is not None else ''
            guardians = s.guardians
            if guardians:
                guardian1 = guardians[0]
                guardian2 = guardians[1] if len(guardians) > 1 else self.blank_guardian
            else:
                guardian1 = guardian2 = self.blank_guardian

            print(f"{s.title:30} {guardian1.title():30} {guardian1.email:30} {guardian1.phone if guardian1.phone else '':12}")
            if address:
                print(f"{"":4} {address:25} {guardian2.title():30} {guardian2.email:30} {str(guardian2.phone):12}")
        print("\n\n")

    def finish(self, data):
        for letter, students in data.students_index:
            print(f"{letter}:")
            for s in sorted(list(students), key=lambda x: x.name):
                print(f"  {s.index_name:30} {s.grade} {s.teacher.class_list_lookup.title()}")
            print("")

class ExcelOutput:

    hyperlink_font = Font(name='Arial', underline='single', size=10, color='3366FF')

    @staticmethod
    def google_width(num):
        return num / 7

    def __init__(self, data):
        self.wb = openpyxl.Workbook()

        self.data = data

        DEFAULT_FONT.name = 'Arial'
        DEFAULT_FONT.size = 10
        medium_border = Side(border_style='medium', color='000000')
        self.medium_border = medium_border
        thin_border = Side(border_style='thin', color='000000')
        self.thin_border = thin_border

        heading = NamedStyle(name='heading')
        heading.alignment = Alignment(horizontal='center', vertical='center')
        heading.font = Font(name='Arial', bold=True, size=12)
        self.wb.add_named_style(heading)

        subheading = NamedStyle(name='subheading')
        subheading.alignment = Alignment(horizontal='center', vertical='center')
        subheading.font = Font(name='Arial', bold=True, size=11)
        self.wb.add_named_style(subheading)

        tableheading = NamedStyle(name='tableheading')
        tableheading.font = Font(name='Arial', bold=True, size=9)
        tableheading.border = Border(top=medium_border, bottom=medium_border)
        self.wb.add_named_style(tableheading)

        tableheadingbegin = NamedStyle(name='tableheadingbegin')
        tableheadingbegin.font = Font(name='Arial', bold=True, size=9)
        tableheadingbegin.border = Border(top=medium_border, bottom=medium_border, left=medium_border)
        self.wb.add_named_style(tableheadingbegin)

        tableheadingend = NamedStyle(name='tableheadingend')
        tableheadingend.font = Font(name='Arial', bold=True, size=9)
        tableheadingend.border = Border(top=medium_border, bottom=medium_border, right=medium_border)
        self.wb.add_named_style(tableheadingend)

        student = NamedStyle(name='student')
        student.font = Font(name='Arial', bold=True, size=11)
        student.border = Border(left=thin_border)
        self.wb.add_named_style(student)

        studentend = NamedStyle(name='studentend')
        studentend.border = Border(right=thin_border)
        self.wb.add_named_style(studentend)

        indexletter = NamedStyle(name='indexletter')
        indexletter.alignment = Alignment(horizontal='center', vertical='bottom')
        indexletter.font = Font(name='Arial', bold=True, size=9)
        self.wb.add_named_style(indexletter)

        indexstudent = NamedStyle(name='indexstudent')
        indexstudent.font = Font(name='Arial', size=10)
        indexstudent.border = Border(top=thin_border, right=thin_border, bottom=thin_border, left=thin_border)
        self.wb.add_named_style(indexstudent)

        self.create_welcome()
        # The links in the TOC are not preserved when sheets exports as PDF
        # self.create_toc()
        self.create_staff()

    def create_welcome(self):
        ws = self.wb.create_sheet(title='Welcome')
        margins = ws.page_margins
        margins.left = margins.right = margins.bottom = 0.25
        margins.top = 0.5

        ws.column_dimensions['A'].width = 97

        ws['A2'] = 'PTA PHONE BOOK'
        ws['A2'].font = Font(name='Arial', bold=True, size=24)
        ws['A2'].alignment = Alignment(horizontal='center')

        ws['A3'] = YEAR
        ws['A3'].font = Font(name='Arial', bold=True, size=14)
        ws['A3'].alignment = Alignment(horizontal='center')

        ws['A6'] = 'WILLIAM HAMMERSCHMIDT SCHOOL'
        ws['A6'].font = Font(name='Arial', bold=True, size=20)
        ws['A6'].alignment = Alignment(horizontal='center')

        ws['A7'] = '617 Hammerschmidt Avenue'
        ws['A7'].font = Font(name='Arial', size=14)
        ws['A7'].alignment = Alignment(horizontal='center')

        ws['A8'] = 'Lombard, IL 60148'
        ws['A8'].font = Font(name='Arial', size=14)
        ws['A8'].alignment = Alignment(horizontal='center')

        ws['A11'] = 'Phone: 630-827-4200    Fax: 630-620-3733'
        ws['A11'].font = Font(name='Arial', size=14)
        ws['A11'].alignment = Alignment(horizontal='center')

        ws['A12'] = 'Voicemail / Attendance: 630-827-4201'
        ws['A12'].font = Font(name='Arial', size=14)
        ws['A12'].alignment = Alignment(horizontal='center')

        big_link = copy(self.hyperlink_font)
        big_link.size = 14

        ws['A13'] = '=hyperlink("https://wh.sd44.org")'
        ws['A13'].font = big_link
        ws['A13'].alignment = Alignment(horizontal='center')

        ws['A15'] = 'School District 44'
        ws['A15'].font = Font(name='Arial', size=14)
        ws['A15'].alignment = Alignment(horizontal='center')

        ws['A16'] = '=hyperlink("https://www.sd44.org")'
        ws['A16'].font = big_link
        ws['A16'].alignment = Alignment(horizontal='center')

        # ws['A19'] = 'Mr. David Danielski'
        # ws['A19'].font = Font(name='Arial', size=14)
        # ws['A19'].alignment = Alignment(horizontal='center')

        # ws['A20'] = 'PRINCIPAL'
        # ws['A20'].font = Font(name='Arial', size=14)
        # ws['A20'].alignment = Alignment(horizontal='center')

        # ws['A22'] = 'Ms. Liz Valdivia'
        # ws['A22'].font = Font(name='Arial', size=14)
        # ws['A22'].alignment = Alignment(horizontal='center')

        # ws['A23'] = 'SECRETARY'
        # ws['A23'].font = Font(name='Arial', size=14)
        # ws['A23'].alignment = Alignment(horizontal='center')

        ws['A33'] = "THIS PTA PHONE BOOK IS FOR PARENT AND STUDENT USE ONLY,\nNOT FOR COMMERCIAL USE."
        ws['A33'].font = Font(name='Arial', size=14)
        ws['A33'].alignment = Alignment(horizontal='center', wrapText=True)

        ws['A35'] = 'This Phone Book is sponsored by the WHS PTA and is issued free, one per member family.'
        ws['A35'].font = Font(name='Arial', size=12)
        ws['A35'].alignment = Alignment(horizontal='center')

    def create_toc(self):
        ws = self.wb.create_sheet(title='Table of Contents')
        margins = ws.page_margins
        margins.left = margins.right = margins.bottom = 0.25
        margins.top = 0.5

        ws.column_dimensions['A'].width = ExcelOutput.google_width(250)
        ws.column_dimensions['B'].width = ExcelOutput.google_width(124)
        ws.column_dimensions['C'].width = ExcelOutput.google_width(35)
        ws.column_dimensions['D'].width = ExcelOutput.google_width(138)

        ws['C1'] = 'Table of Contents'
        ws['C1'].style = 'heading'

        ws['B3'] = 'Resources'
        ws['B3'].style = 'subheading'
        ws['B3'].alignment = Alignment()

        ws['B4'] = 'Staff'
        ws['B4'].hyperlink = Hyperlink(ref="A1", location=f"Staff!A1", target=None)
        ws['B4'].font = Font(underline='single', color='000000')

        ws['B5'] = 'Student Index'
        ws['B5'].hyperlink = Hyperlink(ref="A1", location=f"Student Index!A1", target=None)
        ws['B5'].font = Font(underline='single', color='000000')

        ws['B6'] = 'PTA Board'
        ws['B6'].hyperlink = Hyperlink(ref="A1", location=f"PTA Board!A1", target=None)
        ws['B6'].font = Font(underline='single', color='000000')

        ws['D3'] = 'Class Lists'
        ws['D3'].style = 'subheading'
        ws['D3'].alignment = Alignment()

        index = 4
        for c in data.class_lists:
            ws[f'D{index}'].value = c.title()
            ws[f'D{index}'].hyperlink = Hyperlink(ref="A1", location=f"{c.title()}!A1", target=None)
            ws[f'D{index}'].font = Font(underline='single', color='000000')
            index += 1


    def create_staff(self):
        ws = self.wb.create_sheet(title='Staff')
        margins = ws.page_margins
        margins.left = margins.right = margins.bottom = 0.25
        margins.top = 0.5

        ws.column_dimensions['A'].width = ExcelOutput.google_width(128)
        ws.column_dimensions['B'].width = ExcelOutput.google_width(140)
        ws.column_dimensions['C'].width = ExcelOutput.google_width(35)
        ws.column_dimensions['D'].width = ExcelOutput.google_width(26)
        ws.column_dimensions['E'].width = ExcelOutput.google_width(114)
        ws.column_dimensions['F'].width = ExcelOutput.google_width(33)
        ws.column_dimensions['G'].width = ExcelOutput.google_width(72)
        ws.column_dimensions['H'].width = ExcelOutput.google_width(146)
        ws.column_dimensions['I'].width = ExcelOutput.google_width(34)

        ws.merge_cells('A1:I1')
        ws['A1'] = 'HAMMERSCHMIDT STAFF'
        ws['A1'].style = 'heading'
        ws['A1'].alignment = Alignment(horizontal = 'center', vertical='bottom')

        ws['A3'] = 'OFFICE'
        ws['A3'].font = Font(name='Arial', bold=True, size=10)
        ws['A3'].border = Border(bottom=self.thin_border)
        ws['B3'] = 'EMAIL'
        ws['B3'].font = Font(name='Arial', bold=True, size=10)
        ws['B3'].border = Border(bottom=self.thin_border)
        ws['C3'] = 'EXT.'
        ws['C3'].font = Font(name='Arial', bold=True, size=10)
        ws['C3'].border = Border(bottom=self.thin_border)

        ws['E3'] = 'TEACHER'
        ws['E3'].font = Font(name='Arial', bold=True, size=10)
        ws['E3'].border = Border(bottom=self.thin_border)
        ws['F3'].border = Border(bottom=self.thin_border)
        ws['G3'] = 'GRADE'
        ws['G3'].font = Font(name='Arial', bold=True, size=10)
        ws['G3'].border = Border(bottom=self.thin_border)
        ws['H3'] = 'EMAIL'
        ws['H3'].font = Font(name='Arial', bold=True, size=10)
        ws['H3'].border = Border(bottom=self.thin_border)
        ws['I3'] = 'EXT.'
        ws['I3'].font = Font(name='Arial', bold=True, size=10)
        ws['I3'].border = Border(bottom=self.thin_border)


    def print_class(self, cls):
        ws = self.wb.create_sheet(title=cls.title())
        margins = ws.page_margins
        margins.left = margins.right = margins.bottom = 0.25
        margins.top = 0.5

        ws.merge_cells('A1:E1')
        ws.merge_cells('A2:E2')
        ws.column_dimensions['A'].width = ExcelOutput.google_width(87)
        ws.column_dimensions['B'].width = ExcelOutput.google_width(128)
        ws.column_dimensions['C'].width = ExcelOutput.google_width(150)
        ws.column_dimensions['D'].width = ExcelOutput.google_width(250)
        ws.column_dimensions['E'].width = ExcelOutput.google_width(100)

        ws['A1'] = cls.teacher.title
        ws['A1'].style = 'heading'
        ws['A2'] = f"{cls.grade.pretty()} - {cls.room}"
        ws['A2'].style = 'subheading'

        ws.append([])
        ws.append(['Student', 'Family Address', 'Parent/Guardian', 'Email', 'Phone'])
        ws['A4'].style = 'tableheadingbegin'
        ws['B4'].style = 'tableheading'
        ws['C4'].style = 'tableheading'
        ws['D4'].style = 'tableheading'
        ws['E4'].style = 'tableheadingend'

        idx = 5

        for s in cls.students:
            address = s.address() if s.address() is not None else ''
            guardians = s.guardians if s.guardians else []
            ws.insert_rows(idx=idx)
            ws.cell(row=idx, column=1)
            ws[f'A{idx}'] = s.title
            ws[f'A{idx}'].style = 'student'
            ws[f'E{idx}'].style = 'studentend'
            num_guardians = len(guardians)

            if num_guardians > 0:
                ws[f'C{idx}'] = guardians[0].title()
                ws[f'D{idx}'].value = f'=hyperlink("{guardians[0].email_link()}", "{guardians[0].email}")'
                ws[f'D{idx}'].font = self.hyperlink_font
                ws[f'D{idx}'].alignment = Alignment(wrap_text=True, vertical='center')

                if guardians[0].phone:
                    ws[f'E{idx}'].value = f'=hyperlink("{guardians[0].phone_link()}", "{guardians[0].phone}")'
                ws[f'E{idx}'].font = self.hyperlink_font
                ws[f'E{idx}'].alignment = Alignment(wrap_text=True, vertical='center')

                if num_guardians > 1 or address:
                    idx += 1
                    ws.insert_rows(idx=idx)
                    ws[f'A{idx}'].style = 'student'
                    ws[f'E{idx}'].style = 'studentend'
                    if address:
                        ws[f'B{idx}'] = address
                    if num_guardians > 1:
                        ws[f'C{idx}'] = guardians[1].title()
                        ws[f'D{idx}'].hyperlink = guardians[1].email_link()
                        ws[f'D{idx}'].value = guardians[1].email
                        ws[f'E{idx}'].hyperlink = guardians[1].phone_link()
                        ws[f'E{idx}'].value = guardians[1].phone
            # Put border on bottom
            ws[f'A{idx}'].border = Border(left=self.thin_border, bottom=self.thin_border)
            ws[f'B{idx}'].border = Border(bottom=self.thin_border)
            ws[f'C{idx}'].border = Border(bottom=self.thin_border)
            ws[f'D{idx}'].border = Border(bottom=self.thin_border)
            ws[f'E{idx}'].border = Border(bottom=self.thin_border, right=self.thin_border)

            idx += 1

    def finish(self, data):
        self.wb.remove(self.wb.active)
        self.create_index(data)
        self.create_thank_you_page()
        self.create_pta_board_page()
        self.wb.save(args.output)

    def create_thank_you_page(self):
        ws = self.wb.create_sheet(title='Thank You')
        margins = ws.page_margins
        margins.left = margins.right = 0.7
        margins.top = margins.bottom = 0.75

        ws.column_dimensions['A'].width = ExcelOutput.google_width(245)
        ws.column_dimensions['B'].width = ExcelOutput.google_width(194)
        ws.column_dimensions['C'].width = ExcelOutput.google_width(90)

        ws['B2'] = 'Many thanks to Mr. Hoganson and Graphics Arts Services, Inc.'
        ws['B2'].style = 'subheading'
        ws['B3'] = 'for helping format this PTA phone book!'
        ws['B3'].style = 'subheading'

        ws['B5'] = 'Thank you so much to Mrs. Hoganson for coordinating'
        ws['B5'].style = 'subheading'
        ws['B6'] = 'the student-created artwork for the covers.'
        ws['B6'].style = 'subheading'

        ws['B8'] = 'Great job to the many students who submitted artwork for the covers!'
        ws['B8'].style = 'subheading'
        ws['A10'] = 'Front cover'
        ws['A10'].alignment = Alignment(horizontal='right')
        ws['A11'] = 'Inside front cover'
        ws['A11'].alignment = Alignment(horizontal='right')
        ws['A12'] = 'Inside back cover'
        ws['A12'].alignment = Alignment(horizontal='right')
        ws['A13'] = 'Back cover'
        ws['A13'].alignment = Alignment(horizontal='right')

        ws['B10'] = artists['front'][0]
        ws['B10'].style = 'subheading'
        ws['B10'].alignment = Alignment(horizontal = 'center', vertical='bottom')
        ws['C10'] = artists['front'][1]

        ws['B11'] = artists['front inside'][0]
        ws['B11'].style = 'subheading'
        ws['B11'].alignment = Alignment(horizontal = 'center', vertical='bottom')
        ws['C11'] = artists['front inside'][1]

        ws['B12'] = artists['back inside'][0]
        ws['B12'].style = 'subheading'
        ws['B12'].alignment = Alignment(horizontal = 'center', vertical='bottom')
        ws['C12'] = artists['back inside'][1]

        ws['B13'] = artists['back'][0]
        ws['B13'].style = 'subheading'
        ws['B13'].alignment = Alignment(horizontal = 'center', vertical='bottom')
        ws['C13'] = artists['back'][1]

        img = openpyxl.drawing.image.Image('images/making-a-difference.png')
        img.anchor = 'B15'
        ws['B15'].alignment = Alignment(horizontal='center')
        ws.add_image(img)

        ws['B25'] = 'Want to get involved in the PTA?'
        ws['B25'].font = Font(name='Arial', bold=True, size=14)
        ws['B25'].alignment = Alignment(horizontal='center')
        ws['B27'] = 'There are opportunities year round to help make'
        ws['B27'].font = Font(name='Arial', bold=True, size=14)
        ws['B27'].alignment = Alignment(horizontal='center')
        ws['B28'] = 'Hammerschmidt even more amazing for our kids.'
        ws['B28'].font = Font(name='Arial', bold=True, size=14)
        ws['B28'].alignment = Alignment(horizontal='center')
        ws['B29'] = 'We have options for every parent and every schedule!'
        ws['B29'].font = Font(name='Arial', bold=True, size=14)
        ws['B29'].alignment = Alignment(horizontal='center')

        ws['B31'] = 'Please join us at a monthly PTA meeting'
        ws['B31'].font = Font(name='Arial', bold=True, size=14)
        ws['B31'].alignment = Alignment(horizontal='center')
        ws['B32'] = 'or contact president.whspta@gmail.com'
        ws['B32'].font = Font(name='Arial', bold=True, size=14)
        ws['B32'].alignment = Alignment(horizontal='center')
        ws['B33'] = 'for more information about how you can get involved. '
        ws['B33'].font = Font(name='Arial', bold=True, size=14)
        ws['B33'].alignment = Alignment(horizontal='center')

    def create_pta_board_page(self):
        ws = self.wb.create_sheet(title='PTA Board')
        margins = ws.page_margins
        margins.left = margins.right = 0.25
        margins.top = margins.bottom = 0.25

        ws.column_dimensions['A'].width = 30.7
        ws.column_dimensions['B'].width = 27.8
        ws.column_dimensions['C'].width = 272
        ws.merge_cells('A1:C1')

        ws['B2'] = f'William Hammerschmidt School PTA Board {YEAR}'
        ws['B2'].alignment = Alignment(horizontal='center')
        ws['B2'].font = Font(size=11, bold=True)

    def create_index(self, data):
        ws = self.wb.create_sheet(title='Student Index')
        margins = ws.page_margins
        margins.left = margins.right = 0.15
        margins.top = 0.35
        margins.bottom = 0.25

        ws.merge_cells('A1:G1')
        ws.merge_cells('A2:G2')
        ws['A1'] = 'STUDENT INDEX'
        ws['A1'].style = 'heading'
        ws['A1'].alignment = Alignment(horizontal = 'center', vertical='bottom')
        ws['A1'].font = Font(name='Arial', bold=True, size=11)
        ws['A1'].border = Border(bottom=self.medium_border)
        ws['A2'] = 'Last Name, First Name, Grade, Teacher'
        ws['A2'].style = 'subheading'
        ws['A2'].font = Font(name='Arial', bold=True, size=10)

        name_width = 25
        grade_width = 3.85
        teacher_width = 14.42

        ws.column_dimensions['A'].width = name_width
        ws.column_dimensions['B'].width = grade_width
        ws.column_dimensions['C'].width = teacher_width
        ws.column_dimensions['D'].width = 4.7
        ws.column_dimensions['E'].width = name_width
        ws.column_dimensions['F'].width = grade_width
        ws.column_dimensions['G'].width = teacher_width

        pos = ExcelIndexPositioner(columns=[('A', 'B', 'C'), ('E', 'F', 'G')])

        for letter, students in data.students_index:
            pos.next_letter()

            # print(f"{pos.letter_merge()} : {pos.letter()}: {letter}")
            ws.merge_cells(pos.letter_merge())
            ws[pos.letter()] = letter
            ws[pos.letter()].style = 'indexletter'

            for s in sorted(list(students), key=lambda x: x.name):
                pos.next_student()
                ws[pos.pos(0)] = s.index_name
                ws[pos.pos(0)].style = 'indexstudent'
                ws[pos.pos(1)] = str(s.grade)
                ws[pos.pos(1)].style = 'indexstudent'
                ws[pos.pos(2)] = s.teacher.class_list_lookup.title()
                ws[pos.pos(2)].style = 'indexstudent'
                # print(f"{pos.pos()}: {s.index_name:30} {s.grade} {s.teacher.class_list_lookup.title()}")
            # print("")

        for row_num in range(1, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 12



class ExcelIndexPositioner:
    """
    Produces a position for writing students into the index
    using a two column structure
    """

    def __init__(self, page_height=62, page_buffer=4, columns=['A', 'E'], letter_start_index=1, student_start_index=1):
        self.letter_start_index = letter_start_index
        self.student_start_index = student_start_index
        self.index = 2
        self.page_height = page_height
        self.page_buffer = page_buffer
        self.columns = columns
        self.column_index = 0
        self.page = 0
        self.page_start_letter = True # First page always starts with a letter

    def letter_merge(self):
        return f"{self.columns[self.column_index][0]}{self.index-1}:{self.columns[self.column_index][-1]}{self.index}"

    def letter(self):
        return f"{self.columns[self.column_index][0]}{self.index-1}"

    def pos(self, col=0):
        return f"{self.columns[self.column_index][col]}{self.index}"

    def next_letter(self):
        self.allocate_space(size=3, buffer=4, start_index=self.letter_start_index, is_letter=True)

    def is_last_column(self):
        return self.column_index == len(self.columns) - 1

    def has_enough_space_in_column(self, buffer):
        limit = (self.page+1) * self.page_height - buffer
        # print(f"  {self.index} < {limit}? {self.index < limit}")
        return self.index < limit

    def next_student(self):
        self.allocate_space(size=1, buffer=1, start_index=self.student_start_index)

    def allocate_space(self, size, buffer, start_index, is_letter=False):
        if self.has_enough_space_in_column(buffer=buffer):
            # Enough space, use same column
            self.index += size
        elif self.is_last_column():
            # Move down a page
            self.page += 1
            page_start = self.page * self.page_height + 1
            # print(f"Moving down a page to start index of {start_index} added to {page_start}")
            self.column_index = 0
            self.index = page_start + start_index
            self.page_start_letter = is_letter
        else:
            # Move over a column
            self.column_index = (self.column_index + 1) % len(self.columns)
            column_start = self.page * self.page_height + 1
            self.index = column_start + start_index
            if self.page == 0:
                self.index += 3
            if self.page_start_letter:
                self.index += 1
            # print(f"  shifting a column and moving to start at index {self.index}")

parser = argparse.ArgumentParser(prog='PROG', usage='%(prog)s [options]')
parser.add_argument('--output', help='the output file path')
parser.add_argument('--parent-files', nargs='+', help='the parent directory files')
parser.add_argument('--class-list', help='the class list file')

args = parser.parse_args()

class AllData:
    def __init__(self, class_lists, students):
        self.class_lists = class_lists
        self.students = students

        self.__update_class_list_data()
        self.__create_student_index()

    def __update_class_list_data(self):
        # Replace students in the class list with those from the parent information
        # since it contains more information
        for c in class_lists:
            parent_students = {s: s for s in students if s.teacher == c.teacher}
            class_students = []
            for s in c.students:
                if s in parent_students:
                    class_students.append(parent_students[s])
                    del parent_students[s]
                else:
                    class_students.append(s)
            class_students.sort()

            if parent_students:
                print(f"WARNING: Found the following students in the parent data not listed in the class data for teacher {c.teacher}")
                print("\n".join(str(s) for s in parent_students.keys()))

            c.students = class_students

            # Replace the teacher with the data from the parent information since it has their full name
            c.teacher = class_students[0].teacher

    def __create_student_index(self):
        all_students = []
        for c in class_lists:
            all_students.extend(c.students)

        by_last_name_first_letter = lambda x: x.name[0]
        self.students_index = groupby(sorted(all_students, key=by_last_name_first_letter), key=by_last_name_first_letter)


class_lists = ClassListParser.parse_lists(args.class_list)
students = ParentParser.parse_parent_students(args.parent_files)

data = AllData(class_lists, students)

for output in [ ExcelOutput(data)]:
    for c in data.class_lists:
        output.print_class(c)
    output.finish(data)
