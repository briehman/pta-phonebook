import argparse
import openpyxl
import os
import re
import sys

from openpyxl import Workbook
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import DEFAULT_FONT, Alignment, Border, Font, NamedStyle, Side

class StaffWriter:
    def __init__(self, file):
        self.file = file

    def write(self):
        print(self.file)
        prefix, ext = os.path.splitext(self.file)
        newfile = f"{prefix}-modified{ext}"
        print(newfile)

        wb = openpyxl.load_workbook(self.file)
        sheet = wb.active

        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        DEFAULT_FONT.name = 'Arial'
        DEFAULT_FONT.size = 10

        for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column)):
            if row_index > 100:
                break
            for cell in row:
                new_cell = new_ws.cell(row=row_index+1, column=cell.col_idx)
                match cell.col_idx:
                    case 1:
                        if cell.value:
                            new_cell.value = cell.value.title()
                            new_cell.alignment = Alignment(horizontal='left')
                    case 2:
                        new_cell.value = cell.value
                        new_cell.alignment = Alignment(horizontal='left')
                    case 3:
                        if cell.value:
                            new_cell.value = f"{cell.value.lower()}@sd44.org"
                            new_cell.hyperlink = f"mailto:{cell.value.lower()}@sd44.org"
                            new_cell.font = Font(underline='single', color='000000')
                            new_cell.alignment = Alignment(horizontal='left')
                    case 4:
                        if isinstance(cell.value, (int, float)):
                            new_cell.value = int(cell.value)
                            new_cell.hyperlink = f"https://call.ctrlq.org/1630827{int(cell.value)}"
                            new_cell.font = Font(underline='single', color='000000')
                            new_cell.alignment = Alignment(horizontal='left')



                # print(f"col {cell.col_idx} row {row_index+1} = {new_cell.value}")
                # if cell.col_idx == 0 and cell.value:
                #     new_cell.value = new_cell.value.title()


        print(f"Saving file to {newfile}")
        new_wb.save(newfile)
            # print(row.index())
            # for col in row:
            #     print(col.column)
            # sys.exit(1)
        print("read it")

parser = argparse.ArgumentParser(prog='PROG', usage='%(prog)s [options]')
parser.add_argument('--staff-list', help='the staff spreadsheet')

args = parser.parse_args()

writer = StaffWriter(args.staff_list)
writer.write()
